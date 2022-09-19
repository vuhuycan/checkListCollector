import openpyxl

#------------------------------------------------------------------------------------------------------
# classes definitions:
# we have 3 types of object to deal with here:
# + sheet : provided by openpyxl library
# + cmd : the commands we want to run and paste to excel file
# + WorkDir : the directory that said commands runs on
#-------------------
class WorkDir:
  #class attribute (all instances of class WorkDir has same value)
  #color = wkDirColor

  #instance attribute
  def __init__ (self, name=None, row=None, col=None):
    self.name = name 
    self.row = row 
    self.col = col 

   #self.checkDirProtection()

  #check if the working dir contain the string in dirProtect
  def checkDirProtection(self):
    global wrnCnt
    if dirProtect != None:
        if ( not re.search(dirProtect,self.name) ):
            logger.warning(f'WARNING!! The dir does not contain "{dirProtect}"!')
            wrnCnt = wrnCnt +1
            proceed = False
        else: proceed = True
        while proceed is False:
            usrIn = input('Please type in a new one or ignore (press "Enter" to ignore):')
            if usrIn == '': 
                proceed = True
            elif not  re.search(dirProtect,usrIn):
                logger.warning(f'WARNING!! This dir does not contain "{dirProtect}"!')
            elif not os.path.isdir(usrIn):
                logger.warning("WARNING!! This dir does not exists!")
            else: 
                sheet.cell(self.row,self.col).value = usrIn #update the workdir in checklist
                logger.info('At row {self.row}, org dir has been replaced by user input:')
                logger.info(f'{usrIn}')
                self.name = usrIn
                proceed = True

  def printToLog(self):
      logger.info(f'row:{self.row}:col:{self.col}: {self.name}')


class cmd:
    #class attribute (all instances of class WorkDir has same value)
   #color = cmdColor
   #if enterBash == True: choice = 'y'
   #else: choice = 'no to all'
    
    #instance attribute
    def __init__ (self, name=None, allArgs =None, row=None, col=None, wkDir=None):
        self.name = name #only command name (arg 0)
        self.allArgs = allArgs #command + other args (all arguments)
        self.row = row 
        self.col = col 
        
        self.wkDir = wkDir
        self.isCollectable = True
        
        self.RsltStaRow = self.row + 1 
       #self.RsltStaCol = self.col 
        self.RsltEndRow = None 
        self.RsltEndCol = None 
        
        self.returncode = 0
        self.stdoutList = []
        self.stderrList = []
        
        self.neededLines = 0 #total lines of stdout + stderr
        self.insertRows = 0  #number of excel rows need to insert for this cmd
        self.newRow = None 
        self.newRsltEndRow = None 
        
        self.evalCmd()

    #----------------------------------------------------
    #export returncode, stdout and stderr to screen
    #-------------------
    def printCmdRslt(self):
       #print(self.returncode)
        for line in self.stdoutList:
            print(line.decode('utf-8'))
        for line in self.stderrList:
            print(line.decode('utf-8'))
    #----------------------------------------------------
    #export returncode, stdout and stderr to log file
    #-------------------
    def printCmdRsltToLog(self):
       #logger.info(self.returncode)
        for line in self.stdoutList:
            logger.info(line.decode('utf-8'))
        for line in self.stderrList:
            logger.info(line.decode('utf-8'))
    #----------------------------------------------------
    
    #----------------------------------------------------
    #check inhibited commands
    #For ex: if this cmd is 'vi', we can't execute it, since it halt the whole program
    #-------------------
    def chkInhibitedCmds(self):
        global wrnCnt
        if any(self.name == ref for ref in inhibitCmdList):
            self.isCollectable = False
            logger.warning(f'WARNING!!: row:{self.row}:col:{self.col}: inhibited cmd: {self.name}')
            wrnCnt = wrnCnt +1
        #bash does not have ll cmd, and somehow, csh does not works with subprocess
        #so, replace ll with ls -l:
        if self.name == 'll' :
            self.allArgs = re.sub('^ll ','ls -l ',self.allArgs)
            self.name = 'ls'
    #----------------------------------------------------
    #
    #-------------------
    def delOldRslt(self):
       #col = aCmd.col
       #cmdName = aCmd.name
        
        #now, to make room for new result, delete remain content from old checklist, below the cmd:
        logger.debug('deleting old result')
        rowPtr = self.row +1
        cellColor = sheet.cell(rowPtr,self.col).fill.start_color.index
        while ( cellColor == rsltColor ) or ( cellColor == 'FFFFFF00'):
            if self.isCollectable and noDelForVi is True:
                pass
            else:
                sheet.cell(rowPtr,self.col).value = ''
            rowPtr = rowPtr +1
            cellColor = sheet.cell(rowPtr,self.col).fill.start_color.index
        self.RsltEndRow = rowPtr -1
        
        #calculate how many columns is colored in the "result" range, it is needed later:
        logger.debug('calculating colored cols')
        clredCol = self.col
        if rowPtr != self.row+1: #if result is empty, self.RsltEndCol = self.col, else:
            #check at the row below the command, count how many columns is colored with same color (may not be rsltColor variable):
            while (   sheet.cell(self.row+1,clredCol+1).fill.start_color.index 
                   == sheet.cell(self.row+1,clredCol  ).fill.start_color.index   ):
                clredCol = clredCol +1
        self.RsltEndCol = clredCol
        logger.debug(f'colored column = {self.RsltEndCol}')
    #----------------------------------------------------

    #----------------------------------------------------
    #execute a command, get results
    # input:  command, workingdir of said command
    # output: returncode, stdout, stderr
    #-------------------
    def runCmd(self): #.name,self.wkDir):
        cmd = self.allArgs
        cwd = self.wkDir
        result = subprocess.Popen([cmd],cwd=cwd,shell=True,stdout=subprocess.PIPE,stderr=subprocess.PIPE)
        try: 
            returncode = result.wait(timeout=3)
            stdoutList = result.stdout.readlines()
            stderrList = result.stderr.readlines()
        except subprocess.TimeoutExpired: 
            result.terminate()
            returncode = "timeout"
            stdoutList = []
            stderrList = []
            pass
        self.returncode = returncode
        #get 200 last lines of stdout/err:
        stdoutLen = len(stdoutList)
        stderrLen = len(stderrList)
        if stdoutLen > rsltLim:
          self.stdoutList = stdoutList[-rsltLim:]
        else:
          self.stdoutList = stdoutList
        if stderrLen > rsltLim:
          self.stderrList = stderrList[-rsltLim:]
        else:
          self.stderrList = stderrList
    
    #----------------------------------------------------
    #check if we need to change the command in case return code is not zero
    #-------------------
    def changeCmd(self):
        global wrnCnt

        if self.returncode ==0 or (self.returncode ==1 and (self.name == 'grep' or self.name == 'egrep' )):
            pass

       #if self.choice == 'y' or self.choice == 'n' : 
       #    cmd.choice = input("Returncode not zero. Do you want to enter bash$ to find out why (y/n/yes to all/no to all)?")
       #    logger.debug(f'choice - {self.choice}')
       #
       #if self.choice == 'yes to all' or self.choice == 'y' :
        elif enterBash is True :
            self.printCmdRslt()
            print(f'             Returncode is "{self.returncode}". Entered bash$ to find out why.')
            print ("             when you want to leave bash$ and continue, type \"exit\"")
            print ("!!!NOTES!!!: \"cd\" cannot be used here. Be carefull with it.")
            usrIn = input("bash$ ")
            if ( usrIn == "exit"): pass
            else:
                while (1):
                    self.allArgs = usrIn
                    self.runCmd() 
                    self.printCmdRslt()
                    usrIn = input("bash$ ")
                    if ( usrIn == "exit"): break
                sheet.cell(self.row,self.col).value = self.allArgs #update the command in checklist
                logger.info('The command above is replaced by user input:')
                logger.info(f'row:{self.row}:col:{self.col}: {self.allArgs}')
        
        if self.returncode !=0 :
            self.printCmdRsltToLog()
            logger.warning(f'WARNING!! returncode is "{self.returncode}". There might be a problem with this cmd.')
            wrnCnt = wrnCnt +1
    #----------------------------------------------------
    #when a command is detected, it's initial attributes are eval by this method
    #-------------------
    def evalCmd(self):
        self.chkInhibitedCmds()
        self.delOldRslt()
        #check if this cmd is 'vi' or 'vim', we can't execute this, since it halt the whole program:
        if self.isCollectable is False: return
        #Execute the command, so we have the precious result :      
        logger.info('--------------')
        logger.info(f'row:{self.row}:col:{self.col}: {self.allArgs}')
        self.runCmd() 
        self.changeCmd() 
        self.neededLines = len(self.stdoutList) + len(self.stderrList)
    #----------------------------------------------------

    #----------------------------------------------------
    # In case the new result has more lines than old result,
    # color needed rows below the last row of results range by a fait green:
    #-------------------
    def colorNeededRows(self):
        if self.RsltEndRow < self.row + self.neededLines:
            logger.debug('coloring new rows')
       #    logger.debug(f'row:{aCmd.row}:RsltEndRow:{aCmd.RsltEndRow}')
            for clrow in sheet.iter_rows(min_row= self.RsltEndRow +1, max_row= self.row +self.neededLines, 
                                         min_col= self.col,           max_col= self.RsltEndCol):
                for cell in clrow:
                    cell.fill = copy(sheet.cell(self.RsltStaRow,self.col).fill)
                    cell.font = copy(sheet.cell(self.RsltStaRow,self.col).font)
    #----------------------------------------------------
    # If specified, merge across, so conditional formating in the first line is applied to the whole line:
    #-------------------
    def mergeAcross(self):
        if mergeEna is True :
            logger.debug('merge across')
            for mgrow in range(self.RsltStaRow,self.RsltStaRow +self.neededLines):
                sheet.merge_cells(start_row=mgrow, start_column=self.col, end_row=mgrow, end_column=self.RsltEndCol)
    #----------------------------------------------------
    # now there's enough room, formating is beautiful. It's time to dump those result out:
    #-------------------
    def printRsltToExcel(self):
        logger.debug('writing to excel')
       #self.printCmdRslt()
       #logger.debug(f'row:{aCmd.row}:RsltEndRow:{aCmd.RsltEndRow}')
       #logger.debug(f'row:{aCmd.row}:RsltStaRow:{aCmd.RsltStaRow}:RsltEndRow:{aCmd.RsltEndRow}')
        rowWrPtr = self.RsltStaRow
        for line in self.stdoutList:
            sheet.cell(rowWrPtr,self.col).value = line
            rowWrPtr = rowWrPtr +1
        for line in self.stderrList:
            sheet.cell(rowWrPtr,self.col).value = line
            rowWrPtr = rowWrPtr +1
    #----------------------------------------------------



#----------------------------------------------------
#check if a cell contain a known command, put it in cmdList
# input:  cell object and row/column of an excel cell
#         global vars: knownCmdList, cmdColor, cmdsInARowList (list of cmd object in a row)
# output: True or False, updated cmdsInARowList (appended new cmd obj if necessary)
#-------------------
def isThisCellACmd(cellVal,row,col,wkDir):
    if not isinstance(cellVal,str): return False
    #check if the cell start with a known command in knownCmdList
    for refCmd in knownCmdListRegex: 
        if refCmd['regex'].match(cellVal) : 
            logger.debug(f'found cmd {row}:{col}:{cellVal}')
            #check if the cell color match "cmdColor" :
            cellColor = sheet.cell(row,col).fill.start_color.index
           #print('cellcolor ',cellColor)
            if not ( isinstance(cellColor,str) or isinstance(cellColor,int) ) : return False
            if (cellColor != cmdColor): 
                logger.debug(f'color is not same as defined cmdColor!')
                return False
            #put the found command cell value, row and column to cmdDict:
            cmdObj = cmd(refCmd['cmd'],cellVal,row,col,wkDir)
            cmdsInARowList.append(cmdObj)
            return True
    else:
       #logger.debug("It's not a cmd in known cmd list")
        return False

#----------------------------------------------------

#----------------------------------------------------
#search for all working dir in the sheet
#input(global):  sheet, wkDirList (list of workdir in current sheet)
#output: wkDirList (appended wkDir objs if any)
#-------------------
def searchWkDirs():
    global wrnCnt
    sheet_min_column = sheet.min_column
    logger.info('Start searching for "workng dir" keyword :')
    
    row_iterator = sheet.iter_rows(min_row= sheet.min_row,    max_row= sheet.max_row, 
                                   min_col= sheet_min_column, max_col= 10,
                                   values_only = True) #sheet.max_column) #use 10 for max_col to run faster
    for rowId,row in enumerate(row_iterator,sheet.min_row):
        wkDirKeyCol = None
        for colId,cell in enumerate(row,sheet_min_column):
            cellVal = cell
        #---find cell with keyword 'working dir' (no case):
            if not isinstance(cellVal, str): continue
            if not re.match('working dir',cellVal,re.IGNORECASE) : continue
            logger.debug(f'found wkDirKey at row:{rowId}:col:{colId}')
        
            cellColor = sheet.cell(rowId,colId).fill.start_color.rgb
            if not ( isinstance(cellColor,str) or isinstance(cellColor,int) ) : continue
            if not (cellColor == wkDirColor) : continue
            logger.debug(f'this cell color match wkDirColor!')
            
            wkDirKeyCol = colId
            break
        else: continue
        
        #if above loop breaks, means we found "working dir" keyword at cell with column= wkDirKeyCol
        #So, check in the cells to the right of this cell, if there's a dir there:
        for i in range(wkDirKeyCol, 11): #sheet.max_column+1):
            wkDir = row[i-sheet_min_column]
        #---check if this dir exists in the drive:
            if isinstance(wkDir,str) and os.path.isdir(wkDir) : 
        #-------save the wkDir,row,col to wkDirList
                wkDirObj = WorkDir(wkDir,rowId,i)
                logger.info(f'Found workdir:')
                wkDirObj.printToLog()
                wkDirObj.checkDirProtection()
                wkDirList.append(wkDirObj)
                break
        else: 
            logger.warning(f'WARNING!! Found "Working Dir" key at row: {rowId} but no actual Dir found!')
            wrnCnt = wrnCnt +1
            continue

#this one simpler (not checking for "workdir" key), but slower. Since os.path.isdir runs alot.
def searchWkDirs2():
    logger.info('Start searching for workng dirs:')
    
    row_iterator = sheet.iter_rows(min_row= sheet.min_row,    max_row= sheet.max_row, 
                                   min_col= sheet.min_column, max_col= 10, #sheet.max_column, #use 10 for max_col to run faster
                                  ) # values_only = True ) 
    for rowId,row in enumerate(row_iterator,sheet.min_row):
        for colId,cell in enumerate(row,sheet.min_column):
            wkDir = cell.value
            cellColor = cell.fill.start_color.rgb
            if not isinstance(wkDir,str): continue
        #---check if this cell color is not wkDirColor:
            if not ( isinstance(cellColor,str) or isinstance(cellColor,int) ) : continue
            if cellColor != wkDirColor: continue
        #---check if this dir exists in the drive:
            if not os.path.isdir(wkDir) : continue
        #---save the wkDir,row,col to wkDirList
            wkDirObj = WorkDir(wkDir,rowId,colId)
            logger.info(f'Found workdir:')
            wkDirObj.printToLog()
            wkDirObj.checkDirProtection()
            wkDirList.append(wkDirObj)
            break
#----------------------------------------------------