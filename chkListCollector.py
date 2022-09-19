#!/common/appl/python/python-3.8.0/bin/python3

import subprocess
import openpyxl
import sys
import re
from copy import copy
import argparse
import os
import logging
from operator import attrgetter



#v0.1r2022.04.01 : support multiple commands in one row (main target is for FaultInsert and RepairVerif checklists)
#v0.1r2022.04.02 : fix bug: when there's multiple cmds in one row: print out wrong cmd/result, bad padding/coloring the inserted lines
#v0.1r2022.04.04 : now search for "Working dir" and "command" in arbitrary column, not fixed anymore
#v0.1r2022.07.25 : fix bug that noDelForVi, mergeAcross and noEnterBash do not work.
#v0.1r2022.08.02 : fix bug: cannot regconize command start with space character
#v0.1r2022.08.02b: fix bug: add timeout=3secs to Popen commands, so faulty commands does not halt the program
#v0.1r2022.08.03 : add feature: if dir does not contain a user specified string ( dirProtect ), a warning is raised.
#v2.0r2022.08.05 : change program structure to oop for esier maintenance and readability.
#                  add CLI, auto remove existed output file.
#v2.0r2022.08.06 : fix bugs:     insert rows will not keep merged cells format below the inserted row. 
#                                inserting rows does not update new value for sheet.max_row
#                  add features: logging, dbg and resultLimit switches
#                  optimize:     re.compile to speed up regex for searching commands
#                                avoid using sheet.min_column and sheet.max_column every iters (they slowing things down significantly)
#v2.0r2022.08.06b: fix some minor bugs, optimize log messages format
#v2.0r2022.08.07 : fix some more minor bugs, clean up comment/log messages
#v2.1r2022.08.08 : optimize: change row insertion algorithm, it is now much faster.
#v2.1r2022.08.08b: fix bug of new row insertion algorithm.
#v2.1r2022.08.09 : fix bug: command right below an workdir is collected twice (bug since v2.0r2022.08.06)
#v2.1r2022.08.13 : optimize: replace sheet.cell inside loops by generator iter_rows for smaller memory footprint
#                            put searchWkDirs into a separate method, just for readability
#                            use values_only True for iter_rows so isThisCellACmd runs faster

initMsg1='////////////////////////////////////////////////////////////////'
initMsg2='chkListCollector v2.1r2022.08.13 ,alpha. !!!No right reserved!!!'

#----------------------------------------------------
# sometimes, you want to edit this portion to modify tool behavior, 
# but very rare, so I did not make this part as tool inputs:
#-------------------
#the tool only regconize following commands:
knownCmdList = ('cd', 'ls', 'll', 'vi', 'cat', 'head', 'tail', 'grep', 'egrep', 'sed', 'awk', 'diff', 'vimdiff')
#when the tool meets these one of these commands, it would skip it:
inhibitCmdList = ('vi', 'vimdiff') #, 'cat')

#color code of "working dir":
wkDirColor = 'FFFFC000'
#color code of "command execution":
cmdColor = 4
#color code of "command result":
rsltColor = 9

#----------------------------------------------------
knownCmdListRegex = []
for refCmd in knownCmdList:
    cmdDict = { 'cmd' : refCmd, 'regex' : re.compile('\s*'+refCmd) }
    knownCmdListRegex.append( cmdDict )
#for huge list, tuple is faster than list. This list is small, but it's used alot. So just incase, turn it to tuple:
knownCmdListRegex = tuple(knownCmdListRegex) 

#inhibitCmdListRegex = []
#for refCmd in inhibitCmdList:  inhibitCmdListRegex.append( re.compile('\s*'+refCmd) )

#----------------------------------------------------
# parse input arguments:
#-------------------
# create a parser object
parser = argparse.ArgumentParser(description = "This program Find, Execute all commands and Replace their results in your checklist.")

# add arguments
parser.add_argument("-i","--inFile", nargs = 1, metavar = "file", type = str,
                     help = "Specify your input checklist (xlsx file). \
                             Notes: Only xlsx file is accepted;\
                                    all pictures and drawing objects would be lost in the output file. \
                                    These are restrictions of the openpyxl library.")
parser.add_argument("-o","--outFile", nargs = '?', metavar = "file", type = str,
                     help = "(optional) Specify name for your output (xlsx file only). \
                                        Default is <inputfilename>.Rslt.xlsx") 
parser.add_argument("-l","--logFile", nargs = '?', metavar = "file", type = str,
                     help = "(optional) Specify name for log file. \
                                        Default is <inputfilename>.log.") 

parser.add_argument("-sh","--enterBash", action = 'store_true',
                     help = "(optional) If there's error with a command, this switch allow you to enter bash terminal. \
                                        Then you can try out new commands, the tool'll save it back to the checklist for you.") 
parser.add_argument("-m","--mergeAcross", action = 'store_true',
                     help = "(optional) \"merge across\" (behave same as merge across function of MSExcel) all columns in the command's result.") 
parser.add_argument("-k","--keepViRslt", action = 'store_true',
                     help = "(optional) keep the result of vi command instead of deleting it by default.") 
parser.add_argument("-p","--dirProtector", nargs = '?', metavar = "keyStr", type = str,
                     help = "(optional) If this option is used, the tool peform an additional check: \
                                        Every working dir must contain the \"keyStr\", \
                                        a warning is raised if not sastify.") 
parser.add_argument("-L","--resultLimit", nargs = '?', metavar = "num", type = int,
                     help = "(optional) By default, the program only dump last 200 lines of stdout/stderr to checklist.\
                                        (because of performance/readability reason). Use this switch to change it.") 

parser.add_argument("-dbg","--debug", action = 'store_true',
                     help = "(optional) Dump more information to console/log file.")

# parse the arguments from standard input
if len(sys.argv) == 1 : 
    print('No argument specified! Use -h or --help for usage info.')
    quit()
args = parser.parse_args()

inputfile = args.inFile[0]

if args.outFile is not None:
    outputfile = args.outFile
else: outputfile = inputfile.rsplit('.',1)[0] + '.Rslt.xlsx'

if args.logFile is not None:
    logfile = args.logFile
else: logfile = inputfile.rsplit('.',1)[0] + '.log'

enterBash = args.enterBash
mergeEna = args.mergeAcross
noDelForVi = args.keepViRslt
dirProtect = args.dirProtector
if args.resultLimit is not None:
    rsltLim = args.resultLimit
else: rsltLim = 200

dbg = args.debug

#TODO: build warning class, separate types of warning: dir not protected, cmd return code not 0, cmd result too long... info: added lines, ...
wrnCnt = 0
#----------------------------------------------------

#----------------------------------------------------
# logfile handling:
#-------------------
#logging.basicConfig(filename =logfile, filemode = 'w', level =logging.DEBUG) #encoding ='utf-8', 
logger = logging.getLogger()
stdoutHandler = logging.StreamHandler(sys.stdout)
logfilehandler = logging.FileHandler(logfile, mode = 'w')

if dbg:
  logger.setLevel(logging.DEBUG)
  stdoutHandler.setLevel(logging.DEBUG)
  logfilehandler.setLevel(logging.DEBUG)
else:
  logger.setLevel(logging.INFO)
  stdoutHandler.setLevel(logging.INFO)
  logfilehandler.setLevel(logging.INFO)

logger.addHandler(stdoutHandler)
logger.addHandler(logfilehandler)
#----------------------------------------------------

#------------------------------------------------------------------------------------------------------
# classes definitions:
# we have 3 types of object to deal with here:
# + sheet : provided by openpyxl library
# + cmd : the commands we want to run and paste to excel file
# + WorkDir : the directory that said commands runs on
#-------------------
class WorkDir:
  #class attribute (all instances of class WorkDir has same value)
  #color = wkDirColor

  #instance attribute
  def __init__ (self, name=None, row=None, col=None):
    self.name = name 
    self.row = row 
    self.col = col 

   #self.checkDirProtection()

  #check if the working dir contain the string in dirProtect
  def checkDirProtection(self):
    global wrnCnt
    if dirProtect != None:
        if ( not re.search(dirProtect,self.name) ):
            logger.warning(f'WARNING!! The dir does not contain "{dirProtect}"!')
            wrnCnt = wrnCnt +1
            proceed = False
        else: proceed = True
        while proceed is False:
            usrIn = input('Please type in a new one or ignore (press "Enter" to ignore):')
            if usrIn == '': 
                proceed = True
            elif not  re.search(dirProtect,usrIn):
                logger.warning(f'WARNING!! This dir does not contain "{dirProtect}"!')
            elif not os.path.isdir(usrIn):
                logger.warning("WARNING!! This dir does not exists!")
            else: 
                sheet.cell(self.row,self.col).value = usrIn #update the workdir in checklist
                logger.info('At row {self.row}, org dir has been replaced by user input:')
                logger.info(f'{usrIn}')
                self.name = usrIn
                proceed = True

  def printToLog(self):
      logger.info(f'row:{self.row}:col:{self.col}: {self.name}')


class cmd:
    #class attribute (all instances of class WorkDir has same value)
   #color = cmdColor
   #if enterBash == True: choice = 'y'
   #else: choice = 'no to all'
    
    #instance attribute
    def __init__ (self, name=None, allArgs =None, row=None, col=None, wkDir=None):
        self.name = name #only command name (arg 0)
        self.allArgs = allArgs #command + other args (all arguments)
        self.row = row 
        self.col = col 
        
        self.wkDir = wkDir
        self.isCollectable = True
        
        self.RsltStaRow = self.row + 1 
       #self.RsltStaCol = self.col 
        self.RsltEndRow = None 
        self.RsltEndCol = None 
        
        self.returncode = 0
        self.stdoutList = []
        self.stderrList = []
        
        self.neededLines = 0 #total lines of stdout + stderr
        self.insertRows = 0  #number of excel rows need to insert for this cmd
        self.newRow = None 
        self.newRsltEndRow = None 
        
        self.evalCmd()

    #----------------------------------------------------
    #export returncode, stdout and stderr to screen
    #-------------------
    def printCmdRslt(self):
       #print(self.returncode)
        for line in self.stdoutList:
            print(line.decode('utf-8'))
        for line in self.stderrList:
            print(line.decode('utf-8'))
    #----------------------------------------------------
    #export returncode, stdout and stderr to log file
    #-------------------
    def printCmdRsltToLog(self):
       #logger.info(self.returncode)
        for line in self.stdoutList:
            logger.info(line.decode('utf-8'))
        for line in self.stderrList:
            logger.info(line.decode('utf-8'))
    #----------------------------------------------------
    
    #----------------------------------------------------
    #check inhibited commands
    #For ex: if this cmd is 'vi', we can't execute it, since it halt the whole program
    #-------------------
    def chkInhibitedCmds(self):
        global wrnCnt
        if any(self.name == ref for ref in inhibitCmdList):
            self.isCollectable = False
            logger.warning(f'WARNING!!: row:{self.row}:col:{self.col}: inhibited cmd: {self.name}')
            wrnCnt = wrnCnt +1
        #bash does not have ll cmd, and somehow, csh does not works with subprocess
        #so, replace ll with ls -l:
        if self.name == 'll' :
            self.allArgs = re.sub('^ll ','ls -l ',self.allArgs)
            self.name = 'ls'
    #----------------------------------------------------
    #
    #-------------------
    def delOldRslt(self):
       #col = aCmd.col
       #cmdName = aCmd.name
        
        #now, to make room for new result, delete remain content from old checklist, below the cmd:
        logger.debug('deleting old result')
        rowPtr = self.row +1
        cellColor = sheet.cell(rowPtr,self.col).fill.start_color.index
        while ( cellColor == rsltColor ) or ( cellColor == 'FFFFFF00'):
            if self.isCollectable and noDelForVi is True:
                pass
            else:
                sheet.cell(rowPtr,self.col).value = ''
            rowPtr = rowPtr +1
            cellColor = sheet.cell(rowPtr,self.col).fill.start_color.index
        self.RsltEndRow = rowPtr -1
        
        #calculate how many columns is colored in the "result" range, it is needed later:
        logger.debug('calculating colored cols')
        clredCol = self.col
        if rowPtr != self.row+1: #if result is empty, self.RsltEndCol = self.col, else:
            #check at the row below the command, count how many columns is colored with same color (may not be rsltColor variable):
            while (   sheet.cell(self.row+1,clredCol+1).fill.start_color.index 
                   == sheet.cell(self.row+1,clredCol  ).fill.start_color.index   ):
                clredCol = clredCol +1
        self.RsltEndCol = clredCol
        logger.debug(f'colored column = {self.RsltEndCol}')
    #----------------------------------------------------

    #----------------------------------------------------
    #execute a command, get results
    # input:  command, workingdir of said command
    # output: returncode, stdout, stderr
    #-------------------
    def runCmd(self): #.name,self.wkDir):
        cmd = self.allArgs
        cwd = self.wkDir
        result = subprocess.Popen([cmd],cwd=cwd,shell=True,stdout=subprocess.PIPE,stderr=subprocess.PIPE)
        try: 
            returncode = result.wait(timeout=3)
            stdoutList = result.stdout.readlines()
            stderrList = result.stderr.readlines()
        except subprocess.TimeoutExpired: 
            result.terminate()
            returncode = "timeout"
            stdoutList = []
            stderrList = []
            pass
        self.returncode = returncode
        #get 200 last lines of stdout/err:
        stdoutLen = len(stdoutList)
        stderrLen = len(stderrList)
        if stdoutLen > rsltLim:
          self.stdoutList = stdoutList[-rsltLim:]
        else:
          self.stdoutList = stdoutList
        if stderrLen > rsltLim:
          self.stderrList = stderrList[-rsltLim:]
        else:
          self.stderrList = stderrList
    
    #----------------------------------------------------
    #check if we need to change the command in case return code is not zero
    #-------------------
    def changeCmd(self):
        global wrnCnt

        if self.returncode ==0 or (self.returncode ==1 and (self.name == 'grep' or self.name == 'egrep' )):
            pass

       #if self.choice == 'y' or self.choice == 'n' : 
       #    cmd.choice = input("Returncode not zero. Do you want to enter bash$ to find out why (y/n/yes to all/no to all)?")
       #    logger.debug(f'choice - {self.choice}')
       #
       #if self.choice == 'yes to all' or self.choice == 'y' :
        elif enterBash is True :
            self.printCmdRslt()
            print(f'             Returncode is "{self.returncode}". Entered bash$ to find out why.')
            print ("             when you want to leave bash$ and continue, type \"exit\"")
            print ("!!!NOTES!!!: \"cd\" cannot be used here. Be carefull with it.")
            usrIn = input("bash$ ")
            if ( usrIn == "exit"): pass
            else:
                while (1):
                    self.allArgs = usrIn
                    self.runCmd() 
                    self.printCmdRslt()
                    usrIn = input("bash$ ")
                    if ( usrIn == "exit"): break
                sheet.cell(self.row,self.col).value = self.allArgs #update the command in checklist
                logger.info('The command above is replaced by user input:')
                logger.info(f'row:{self.row}:col:{self.col}: {self.allArgs}')
        
        if self.returncode !=0 :
            self.printCmdRsltToLog()
            logger.warning(f'WARNING!! returncode is "{self.returncode}". There might be a problem with this cmd.')
            wrnCnt = wrnCnt +1
    #----------------------------------------------------
    #when a command is detected, it's initial attributes are eval by this method
    #-------------------
    def evalCmd(self):
        self.chkInhibitedCmds()
        self.delOldRslt()
        #check if this cmd is 'vi' or 'vim', we can't execute this, since it halt the whole program:
        if self.isCollectable is False: return
        #Execute the command, so we have the precious result :      
        logger.info('--------------')
        logger.info(f'row:{self.row}:col:{self.col}: {self.allArgs}')
        self.runCmd() 
        self.changeCmd() 
        self.neededLines = len(self.stdoutList) + len(self.stderrList)
    #----------------------------------------------------

    #----------------------------------------------------
    # In case the new result has more lines than old result,
    # color needed rows below the last row of results range by a fait green:
    #-------------------
    def colorNeededRows(self):
        if self.RsltEndRow < self.row + self.neededLines:
            logger.debug('coloring new rows')
       #    logger.debug(f'row:{aCmd.row}:RsltEndRow:{aCmd.RsltEndRow}')
            for clrow in sheet.iter_rows(min_row= self.RsltEndRow +1, max_row= self.row +self.neededLines, 
                                         min_col= self.col,           max_col= self.RsltEndCol):
                for cell in clrow:
                    cell.fill = copy(sheet.cell(self.RsltStaRow,self.col).fill)
                    cell.font = copy(sheet.cell(self.RsltStaRow,self.col).font)
    #----------------------------------------------------
    # If specified, merge across, so conditional formating in the first line is applied to the whole line:
    #-------------------
    def mergeAcross(self):
        if mergeEna is True :
            logger.debug('merge across')
            for mgrow in range(self.RsltStaRow,self.RsltStaRow +self.neededLines):
                sheet.merge_cells(start_row=mgrow, start_column=self.col, end_row=mgrow, end_column=self.RsltEndCol)
    #----------------------------------------------------
    # now there's enough room, formating is beautiful. It's time to dump those result out:
    #-------------------
    def printRsltToExcel(self):
        logger.debug('writing to excel')
       #self.printCmdRslt()
       #logger.debug(f'row:{aCmd.row}:RsltEndRow:{aCmd.RsltEndRow}')
       #logger.debug(f'row:{aCmd.row}:RsltStaRow:{aCmd.RsltStaRow}:RsltEndRow:{aCmd.RsltEndRow}')
        rowWrPtr = self.RsltStaRow
        for line in self.stdoutList:
            sheet.cell(rowWrPtr,self.col).value = line
            rowWrPtr = rowWrPtr +1
        for line in self.stderrList:
            sheet.cell(rowWrPtr,self.col).value = line
            rowWrPtr = rowWrPtr +1
    #----------------------------------------------------



#----------------------------------------------------
#check if a cell contain a known command, put it in cmdList
# input:  cell object and row/column of an excel cell
#         global vars: knownCmdList, cmdColor, cmdsInARowList (list of cmd object in a row)
# output: True or False, updated cmdsInARowList (appended new cmd obj if necessary)
#-------------------
def isThisCellACmd(cellVal,row,col,wkDir):
    if not isinstance(cellVal,str): return False
    #check if the cell start with a known command in knownCmdList
    for refCmd in knownCmdListRegex: 
        if refCmd['regex'].match(cellVal) : 
            logger.debug(f'found cmd {row}:{col}:{cellVal}')
            #check if the cell color match "cmdColor" :
            cellColor = sheet.cell(row,col).fill.start_color.index
           #print('cellcolor ',cellColor)
            if not ( isinstance(cellColor,str) or isinstance(cellColor,int) ) : return False
            if (cellColor != cmdColor): 
                logger.debug(f'color is not same as defined cmdColor!')
                return False
            #put the found command cell value, row and column to cmdDict:
            cmdObj = cmd(refCmd['cmd'],cellVal,row,col,wkDir)
            cmdsInARowList.append(cmdObj)
            return True
    else:
       #logger.debug("It's not a cmd in known cmd list")
        return False

#----------------------------------------------------

#----------------------------------------------------
#search for all working dir in the sheet
#input(global):  sheet, wkDirList (list of workdir in current sheet)
#output: wkDirList (appended wkDir objs if any)
#-------------------
def searchWkDirs():
    global wrnCnt
    sheet_min_column = sheet.min_column
    logger.info('Start searching for "workng dir" keyword :')
    
    row_iterator = sheet.iter_rows(min_row= sheet.min_row,    max_row= sheet.max_row, 
                                   min_col= sheet_min_column, max_col= 10,
                                   values_only = True) #sheet.max_column) #use 10 for max_col to run faster
    for rowId,row in enumerate(row_iterator,sheet.min_row):
        wkDirKeyCol = None
        for colId,cell in enumerate(row,sheet_min_column):
            cellVal = cell
        #---find cell with keyword 'working dir' (no case):
            if not isinstance(cellVal, str): continue
            if not re.match('working dir',cellVal,re.IGNORECASE) : continue
            logger.debug(f'found wkDirKey at row:{rowId}:col:{colId}')
        
            cellColor = sheet.cell(rowId,colId).fill.start_color.rgb
            if not ( isinstance(cellColor,str) or isinstance(cellColor,int) ) : continue
            if not (cellColor == wkDirColor) : continue
            logger.debug(f'this cell color match wkDirColor!')
            
            wkDirKeyCol = colId
            break
        else: continue
        
        #if above loop breaks, means we found "working dir" keyword at cell with column= wkDirKeyCol
        #So, check in the cells to the right of this cell, if there's a dir there:
        for i in range(wkDirKeyCol, 11): #sheet.max_column+1):
            wkDir = row[i-sheet_min_column]
        #---check if this dir exists in the drive:
            if isinstance(wkDir,str) and os.path.isdir(wkDir) : 
        #-------save the wkDir,row,col to wkDirList
                wkDirObj = WorkDir(wkDir,rowId,i)
                logger.info(f'Found workdir:')
                wkDirObj.printToLog()
                wkDirObj.checkDirProtection()
                wkDirList.append(wkDirObj)
                break
        else: 
            logger.warning(f'WARNING!! Found "Working Dir" key at row: {rowId} but no actual Dir found!')
            wrnCnt = wrnCnt +1
            continue

#this one simpler (not checking for "workdir" key), but slower. Since os.path.isdir runs alot.
def searchWkDirs2():
    logger.info('Start searching for workng dirs:')
    
    row_iterator = sheet.iter_rows(min_row= sheet.min_row,    max_row= sheet.max_row, 
                                   min_col= sheet.min_column, max_col= 10, #sheet.max_column, #use 10 for max_col to run faster
                                  ) # values_only = True ) 
    for rowId,row in enumerate(row_iterator,sheet.min_row):
        for colId,cell in enumerate(row,sheet.min_column):
            wkDir = cell.value
            cellColor = cell.fill.start_color.rgb
            if not isinstance(wkDir,str): continue
        #---check if this cell color is not wkDirColor:
            if not ( isinstance(cellColor,str) or isinstance(cellColor,int) ) : continue
            if cellColor != wkDirColor: continue
        #---check if this dir exists in the drive:
            if not os.path.isdir(wkDir) : continue
        #---save the wkDir,row,col to wkDirList
            wkDirObj = WorkDir(wkDir,rowId,colId)
            logger.info(f'Found workdir:')
            wkDirObj.printToLog()
            wkDirObj.checkDirProtection()
            wkDirList.append(wkDirObj)
            break
#----------------------------------------------------
    
    
#------------------------------------------------------------------------------------------------------
# MAIN PRGRAM START HERE
#-------------------
logger.info(initMsg1)
logger.info(initMsg2)
logger.info(initMsg1)
logger.info('')



#open file, check output:
workbk = openpyxl.load_workbook(filename = inputfile)
logger.info(f'Opened {inputfile}')
logger.info('')
if os.path.exists(outputfile):
    os.remove(outputfile)
    logger.info(f'WARNING!! The file {outputfile} exists. It is deleted.')

#get all worksheet in the file:
sheetNameList = workbk.sheetnames

for i in sheetNameList:
    sheet = workbk[i]
    logger.info(f'OPENED SHEET {i}')
    logger.info('')

    wkDirList = [] # <----- THE LIST OF WORKING DIR IN CURRENT SHEET.
    sheet_min_column = sheet.min_column
    sheet_max_column = sheet.max_column
    #----------------------------------------------------
    #search for all working dir in the sheet
    #input(global):  sheet, wkDirList (list of workdir in current sheet)
    #output: wkDirList (appended wkDir objs if any)
    #-------------------
    searchWkDirs()
    #----------------------------------------------------
    wkDirList.append(WorkDir('dummy',sheet.max_row,sheet.max_column))
    #----------------------------------------------------
    logger.info('')
    logger.info('')
    logger.info('Finished looking for working dir in this sheet.')
    if not wkDirList:
        logger.info('No working dir found in it.')
        continue
    logger.debug('wkDirList is:')
    logger.debug(f'{wkDirList}')
    logger.info('')
    logger.info('')
    logger.info('')
    
    cmdList = []
    totalInsertRows = 0

    for i in range(len(wkDirList)-1):#for all found "working dir":
        logger.info('-----------------------------------------------------------')
        logger.info(f'---COLLECTING RESULT IN {wkDirList[i].name} ----:')
        logger.info('')
        logger.info('')
        #----------------------------------------------------
        #for each row between two "working dir", perform:
        #  search for all cmds on one row (execute it in bash right after the cmd object is created)
        #  then check if the result exceed the available space 
        #       + aCmd.EndRsltRow - aCmd.row = available
        #       + aCmd.neededLines           = the amount needed
        #       + aCmd.insertRows            = number of rows will be inserted below aCmd.EndRsltRow
        #  if so, we need to insert rows where needed.
        #
        #TODO write better method, because current algo only works if all cmds in "cmds InARow List" are in on one row
        #-------------------
        row_iterator = sheet.iter_rows(min_row= wkDirList[i].row+1, max_row= wkDirList[i+1].row-1, 
                                       min_col= sheet_min_column,   max_col= sheet_max_column,
                                       values_only = True )
        for rowId,row in enumerate(row_iterator,wkDirList[i].row+1):
           #logger.debug(f'searching for cmds in row {rowId}:')
            cmdsInARowList = []
            #----------------------------------------------------
            # look for known commands
            #-------------------
            foundCmd = False
            for colId,cell in enumerate(row,sheet_min_column):
                rslt = isThisCellACmd(cell,rowId,colId,wkDirList[i].name)
                if rslt is True: foundCmd = True
            if foundCmd == False: continue
            #finished looking for commands in a row. 
            #----------------------------------------------------
            
            cmdList.extend(cmdsInARowList)
           #cmdList.append(cmdsInARowList)
            
            #----------------------------------------------------
            #results of all cmds in current row are now saved in cmds in a row List (cmdsInARowList)
            # In each row, only the command with longest result (biggest aCmd.neededLines) is targeted for row insertion.
            # Other commands in the same rows - with less demanding space, 
            # don't need row insertion anymore, as long as there enough space for the longest one.
            #  So, only one aCmd.insertRows (the one with biggest [neededLines -(EndRsltRow - row)]) 
            #      contain actual aCmd.insertRows value.
            #  The rest aCmd.insertRows is zero.
            #  totalInsertRows contain all rows need to insert in the current sheet. We would need it later.
            #-------------------
            logger.debug('calculating new coordinate of cmds')
            for aCmd in cmdsInARowList:
                aCmd.newRow = aCmd.row + totalInsertRows
                aCmd.newRsltEndRow = aCmd.RsltEndRow + totalInsertRows

            logger.debug('start evaluating row insertion')
            maxNeededLines = max([cmdDict.neededLines for cmdDict in cmdsInARowList])
            maxEndClrRow = max([cmdDict.RsltEndRow for cmdDict in cmdsInARowList]) 
            cmdObj_maxEndClrRow = max(cmdsInARowList,key = attrgetter('RsltEndRow'))
            cmdObj_maxEndClrRow.insertRows = max( 0, maxNeededLines - (maxEndClrRow -rowId) )
            totalInsertRows = totalInsertRows + cmdObj_maxEndClrRow.insertRows
            logger.debug('end evaluating row insertion')
            #----------------------------------------------------
        #end executing in all rows of a "working dir"
        #----------------------------------------------------
    #end executing in a sheet
    #----------------------------------------------------

    logger.info('')
    logger.info("-----------------------------------------------------------")
    logger.info(f'End executing all commands in this sheet. ')
    logger.info('')
    logger.info('Inserting new rows.')
    logger.info('')

    farthest_row  = sheet.max_row
    farthest_coor = sheet.cell(farthest_row,sheet_max_column).coordinate
    
    #----------------------------------------------------
    # Now, for all commands from bottom to top (of the sheet),
    # we shift the space below it ( below aCmd.RsltEndRow ) down 
    #    by an amount (not aCmd.insertRows, but totalInsertRows).
    # 
    # But why from bottom to top?
    # -> it's for performance reason. 
    #  - Each time new rows are added, 
    #    openpyxl spend a large amount of time moving a lot of cells down.
    #    So we want to move cells around as little as possible.
    #  - This is achieved by the variable totalInsertRows.
    #    Each command is moved down by totalInsertRows,
    #    but in each iter, totalInsertRows decreased, = previos totalInsertRows - previos aCmd.insertRows
    #-------------------
    cmdList.sort(key=attrgetter('RsltEndRow')) # this fix works, but result is ugly
    for aCmd in reversed(cmdList):
        logger.debug('')
        logger.debug(f'row:{aCmd.row}:col:{aCmd.col}: CMD: {aCmd.allArgs}')
        logger.debug(f'avail:{aCmd.RsltEndRow-aCmd.row}:need:{aCmd.neededLines}:insert:{aCmd.insertRows}')

        if aCmd.insertRows :
        #----------------------------------------------------
        #---get the move range below the current command:
            curr_coor = sheet.cell(aCmd.RsltEndRow +1,sheet_min_column).coordinate
            toMoveRange = curr_coor+':'+farthest_coor
           #logger.debug(f'curr_coor:{curr_coor}')
           #logger.debug(f'farthest_coor:{farthest_coor}:farthest_row:{farthest_row}')
        #---move range
            logger.info(f'Moving range "{toMoveRange}" down {totalInsertRows} row(s).')
            sheet.move_range(toMoveRange,rows=totalInsertRows)
        #---preseve merged cells format before adding rows
            logger.debug('preseving merged cells format')
            for merged_cells in sheet.merged_cells.ranges:
                if merged_cells.min_row > aCmd.RsltEndRow and merged_cells.max_row <= farthest_row :
                    merged_cells.shift(0,totalInsertRows)

        #----------------------------------------------------
        #---recalculate totalInsertRows
            totalInsertRows = totalInsertRows - aCmd.insertRows
        #---get coordinate of the farthest cell in move range of the next command:
            farthest_coor = sheet.cell(aCmd.RsltEndRow,sheet_max_column).coordinate
            farthest_row  = aCmd.RsltEndRow

    #----------------------------------------------------
    #Last but not least,
    #print out the stdout and stderr to the excel file:
    #-------------------
    logger.debug('')
    logger.debug('----------------------')
    logger.info(f'Row insertion done. Dumping results back to the sheet.')
    logger.debug('---------')
    logger.debug('')
    for aCmd in cmdList:
        logger.debug('')
        logger.debug(f'row:{aCmd.row}:col:{aCmd.col}: CMD: {aCmd.name}')
        logger.debug(f'avail:{aCmd.RsltEndRow-aCmd.row}:need:{aCmd.neededLines}:insert:{aCmd.insertRows}')
        
        # recalculate new rows of aCmd, after insert rows
        logger.debug('recalibrating rows:')
        aCmd.row = aCmd.newRow 
        aCmd.RsltStaRow = aCmd.row + 1
        aCmd.RsltEndRow = aCmd.newRsltEndRow 
        logger.debug(f'row:{aCmd.row}:RsltStaRow:{aCmd.RsltStaRow}:RsltEndRow:{aCmd.RsltEndRow}')

        # color new rows by a fait green:
        aCmd.colorNeededRows()
        # If specified, merge across, so conditional formating in the first line is applied to the whole line:
        aCmd.mergeAcross()
        # now there's enough room, formating is beautiful. It's time to dump those result out:
        aCmd.printRsltToExcel()
        #----------------------------------------------------
    #end collecting in a sheet
    logger.info('')
    logger.info("-----------------------------------------------------------")
    logger.info('')
    logger.info('')
#end whole workbook
logger.info("Completed collecting this checklist.")
logger.info("-----------------------------------------------------------")
logger.info('')
logger.info("-----------------------------------------------------------")
logger.info(f'{wrnCnt} WARNING(S)')
logger.info("-----------------------------------------------------------")
logger.info('')

logger.info(f'Saving to file {outputfile}')
workbk.save(outputfile)
logger.info(f'Output file saved.')
logger.info('')
logger.info(f'Log file is {logfile}.')

logger.info('')
logger.info('')
logger.info("Wish you all the best with the project!")
logger.info("-----------------------------------------------------------")
logfilehandler.close()
